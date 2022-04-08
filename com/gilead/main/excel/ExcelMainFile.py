'''
Created on Oct 23, 2018

@author: bkotamahanti

'''

from definitions import EXCEL_PROCESS
from openpyxl.compat import range
import os
from selenium import webdriver
from openpyxl.reader.excel import load_workbook
from openpyxl.workbook.workbook import Workbook
import pyautogui

class ExcelMainClass():
    __process = EXCEL_PROCESS
    
    def __init__(self, utilsRef, loggerRef):
        self.utils = utilsRef
        self.logger = loggerRef
        self.driver = ""
        self.utils.killProcess(self.__process)
        
    def killProcess(self):
        self.utils.killProcess(self.__process)
        
    def createExcelSheet(self, fileName):
        wb = Workbook()
        ws1 = wb.active
        '''Create Numbers sheet '''
        ws1.title = "Numbers"
        for row in range(1, 40):
            ws1.append(range(10))
        '''Create Names sheet '''
        ws2 = wb.create_sheet(title="Names")
        ws2['A1']="Hello"
        ws2['B1']="World!!"
        ws2['A2']="Hello"
        ws2['B2']="World!!"
        ws2['A3']="Hello"
        ws2['B3']="World!!"
        wb.save(filename = fileName)
    
    def isFileExists(self, filename):
        if os.path.exists(filename):
            self.logger.info("File "+str(filename)+" exists")
            return True
        else:
            self.logger.info("File "+str(filename)+" doesn't exist")
            return False 
    
    def loadWorkBook(self, fileName):
        wb = load_workbook(filename = fileName)
        return wb
    
    def getExcelSheetNames(self, fileName):
        wb = self.loadWorkBook(fileName)
        return wb.sheetnames
    
    def enterDataInWorkSheet(self, worksheet):
        parent_element = self.driver.find_element_by_class_name("ExcelGrid")
        excel_data_element = parent_element.find_element_by_name("MSExcelDocument")
        sheet_element = excel_data_element.find_element_by_name("Sheet "+worksheet)
#         for i in range(1,5):
        row = "A"+str(1)
        sheet_element.find_element_by_name(row).click()
        sheet_element.find_element_by_name(row).clear()
        sheet_element.find_element_by_name(row).send_keys(222)
        
    def clickExcelCloseButton(self):
        parent_element = self.driver.find_element_by_class_name("NetUIOfficeCaption")
        elements = parent_element.find_elements_by_class_name("NetUInetpane")
        for element in elements:
            if "Close" == element.get_attribute("Name"):
                element.click()

    def clickExcelDontSaveMenuOption(self):
        pyautogui.hotkey('alt','f4')

    def getExcelPromptDialogElement(self):
        parent_element = self.driver.find_element_by_class_name("XLMAIN")
        dialog_prompt_element = parent_element.find_element_by_class_name("NUIDialog")
        return dialog_prompt_element
        
    def getWorkSheet(self, fileName, sheetName):
        wb = self.loadWorkBook(fileName)
        if sheetName in wb.sheetnames:
            return wb[sheetName]
    
    def isExcelElementExist(self, webelement, **property1):
        return self.utils.isElementPresent( webelement,  **property1)
    
    def isExcelProcessExist(self):
        return self.utils.isProcessExist(self.__process)
    
    def connectExcelToWebDriver(self):
        self.driver = webdriver.Remote(command_executor="http://localhost:9999",
                                       desired_capabilities={
#                                            'app': r'C:\Program Files (x86)\Microsoft Office\Office16\EXCEL.EXE',
                                           'app': r'C:\Program Files (x86)\Microsoft Office\root\Office16\EXCEL.EXE',
                                           'debugConnectToRunningApp': 'true',
                                           "args": "-port 345"
                                        })
    def clickOKOnExcelOpenDialog(self):
        open_dialog_element = self.getExcelPopUpWindowElement()
        open_dialog_element.find_element_by_name("OK").click()   
    
    def openExcel(self, filePath):
        self.utils.openApplication(filePath) 
    
    def clickCancelOnExcelOpenDialog(self):
        browse_window_element = self.driver.find_element_by_class_name("#32770")
        elements =  browse_window_element.find_elements_by_class_name("Button")
        for element in elements:
            if element.get_attribute("Name") == "Cancel":
                element.click()
                
    def isExcelElementTextExist(self, webelement, text , **property1):
        return self.utils.isElementTextPresent(webelement, text , **property1)
    
    def clickExceltAppBackButton(self):
        parent_element = self.driver.find_element_by_class_name("XLMAIN") 
        parent_element.find_element_by_name("Back").click()
    
    def clickCloseButtonOfExcelAppWhenNoPresentation(self):
        parent_element = self.driver.find_element_by_class_name("XLMAIN")
        parent_element.find_element_by_name("Close").click()
    
    
    def clickBlankExcelWorkbook(self):
        parent_element = self.driver.find_element_by_class_name("XLMAIN")
        parent_element.find_element_by_name("Blank workbook").click()
    
    def clickExcelOpenMenuOption(self):
        pyautogui.hotkey('ctrl','o')
    
    def enterFileNameInOpenBrowseWindow(self, fileName):
        browse_window_element = self.driver.find_element_by_class_name("#32770")
        browse_window_element.find_element_by_class_name("Edit").click()
        browse_window_element.find_element_by_class_name("Edit").send_keys(fileName)
    
    def clickOpenButtonOfOpenBrowseWindow(self):
        browse_window_element = self.driver.find_element_by_class_name("#32770")
        elements =  browse_window_element.find_elements_by_class_name("Button")
        for element in elements:
            if element.get_attribute("Name") == "Open":
                element.click()
    
    def clickOpenBrowseButton(self):
        parent_element = self.driver.find_element_by_class_name("XLMAIN")
        backstage_element= parent_element.find_element_by_class_name("NetUIFullpageUIWindow").find_element_by_name("Backstage view")
        open_grp_element = backstage_element.find_element_by_class_name("NetUIElement")
        open_grp_element.find_element_by_class_name("NetUISlabContainer").find_element_by_name("Browse").click()
    
    def getExcelPopUpWindowElement(self):
        parent_element = self.driver.find_element_by_name("Book1 - Excel")
        open_dialog_element = parent_element.find_element_by_class_name("#32770")
        return open_dialog_element